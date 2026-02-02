from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


class ExcelExportService:
    """Export dataset to Excel format"""
    
    def generate_excel(self, dataset):
        """
        Create Excel file with dataset info and records
        Returns: BytesIO buffer with Excel content
        """
        wb = Workbook()
        
        # summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # header styling
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # add summary info
        ws_summary['A1'] = "Dataset Summary"
        ws_summary['A1'].font = Font(size=14, bold=True)
        
        ws_summary['A3'] = "Filename"
        ws_summary['B3'] = dataset.filename
        ws_summary['A4'] = "Upload Date"
        ws_summary['B4'] = dataset.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
        ws_summary['A5'] = "Total Equipment"
        ws_summary['B5'] = dataset.total_equipment
        
        ws_summary['A7'] = "Statistics"
        ws_summary['A7'].font = Font(bold=True)
        ws_summary['A8'] = "Average Flowrate"
        ws_summary['B8'] = round(dataset.avg_flowrate, 2)
        ws_summary['A9'] = "Average Pressure"
        ws_summary['B9'] = round(dataset.avg_pressure, 2)
        ws_summary['A10'] = "Average Temperature"
        ws_summary['B10'] = round(dataset.avg_temperature, 2)
        
        # type distribution
        ws_summary['A12'] = "Equipment Type Distribution"
        ws_summary['A12'].font = Font(bold=True)
        row = 13
        for eq_type, count in dataset.type_distribution.items():
            ws_summary[f'A{row}'] = eq_type
            ws_summary[f'B{row}'] = count
            row += 1
        
        # data sheet
        ws_data = wb.create_sheet("Equipment Data")
        
        # headers
        headers = ['Equipment Name', 'Type', 'Flowrate', 'Pressure', 'Temperature']
        for col, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # data rows
        records = dataset.records.all()
        for row_idx, record in enumerate(records, start=2):
            ws_data.cell(row=row_idx, column=1, value=record.equipment_name)
            ws_data.cell(row=row_idx, column=2, value=record.equipment_type)
            ws_data.cell(row=row_idx, column=3, value=record.flowrate)
            ws_data.cell(row=row_idx, column=4, value=record.pressure)
            ws_data.cell(row=row_idx, column=5, value=record.temperature)
        
        # adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_data.column_dimensions[col].width = 20
        
        # save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
